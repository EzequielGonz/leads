# -*- coding: utf-8 -*-
"""Excel/CSV reader optimized for low memory usage.

For .xlsx/.xlsm: uses openpyxl read_only=True (streams rows, no full load).
For .csv: uses csv module (streams rows).
For .xls: falls back to pandas (old format, unavoidable).
"""

import csv
import io
import math
import os
import re
from datetime import date, datetime

import chardet

_HEADER_KEYWORDS = [
    "nombre", "apellido", "name", "dni", "documento",
    "telefono", "teléfono", "celular", "phone",
    "email", "correo", "mail", "instagram", "linkedin",
    "web", "sitio", "ubicacion", "ubicación", "ciudad",
    "pais", "país", "provincia", "localidad",
    "direccion", "dirección", "domicilio", "fecha",
    "sexo", "genero", "género", "bio", "descripcion",
    "descripción", "tipo", "categoria", "categoría",
    "empresa", "abogado", "contador", "medico", "médico",
    "doctor", "art", "siniestro", "expediente", "cuit",
    "cuil", "edad", "profesion", "profesión", "ocupacion",
    "ocupación", "estado", "observacion", "observación",
    "detalle", "referencia", "comentario",
]

_DATE_RE = re.compile(r"^\d{1,2}[/-]\d{1,2}[/-]\d{2,4}$")
_DIGITS_RE = re.compile(r"^\d+$")


def _sanitize_value(val):
    if val is None:
        return ""
    if isinstance(val, float) and math.isnan(val):
        return ""
    if isinstance(val, str):
        stripped = val.strip()
        return stripped
    try:
        import pandas as pd
        if isinstance(val, pd.Timestamp):
            return str(val)
    except ImportError:
        pass
    if isinstance(val, (datetime, date)):
        return str(val)
    if isinstance(val, float):
        return str(int(val)) if val.is_integer() else str(val)
    if isinstance(val, bool):
        return str(val)
    if isinstance(val, int):
        return str(val)
    return str(val)


def _sanitize_cell(val):
    """Convert openpyxl cell value to string."""
    if val is None:
        return ""
    if isinstance(val, float) and math.isnan(val):
        return ""
    if isinstance(val, (datetime, date)):
        return str(val)
    if isinstance(val, float):
        return str(int(val)) if val.is_integer() else str(val)
    if isinstance(val, bool):
        return str(val)
    if isinstance(val, int):
        return str(val)
    s = str(val).strip()
    return s


def _looks_like_data_cell(value):
    if value is None:
        return False
    if isinstance(value, float) and math.isnan(value):
        return False
    if isinstance(value, (int, float, datetime, date)):
        return True
    text = str(value).strip()
    if not text:
        return False
    if _DIGITS_RE.match(text) and len(text) >= 4:
        return True
    if _DATE_RE.match(text):
        return True
    return False


def _looks_like_header_row(first_row):
    values = list(first_row)
    total = max(1, len(values))
    data_cells = sum(1 for v in values if _looks_like_data_cell(v))
    keyword_cells = sum(
        1 for v in values
        if isinstance(v, str) and any(k in v.lower() for k in _HEADER_KEYWORDS)
    )
    data_fraction = data_cells / total
    keyword_fraction = keyword_cells / total
    if data_fraction == 0:
        return True
    return keyword_fraction >= 0.5


def get_sheet_names(file_path):
    ext = os.path.splitext(file_path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True, data_only=True)
        names = wb.sheetnames
        wb.close()
        return names
    if ext == ".xls":
        try:
            import xlrd
            wb = xlrd.open_workbook(file_path)
            return wb.sheet_names()
        except Exception:
            return None
    return None


# ---------------------------------------------------------------------------
# Streaming readers — never load entire file into memory
# ---------------------------------------------------------------------------

def _stream_xlsx(file_path, sheet_name=None):
    """Stream rows from .xlsx/.xlsm using openpyxl read_only mode.

    Yields (columns, row_dicts_iterator) where row_dicts_iterator yields
    dicts one at a time.
    """
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True, data_only=True)
    if sheet_name:
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    else:
        ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)

    # Read first row to detect headers
    try:
        first_row = next(rows_iter)
    except StopIteration:
        wb.close()
        return

    first_row_clean = [_sanitize_cell(c) for c in first_row]
    if _looks_like_header_row(first_row_clean):
        columns = []
        for i, cell in enumerate(first_row_clean):
            label = cell.strip() if cell else ""
            columns.append(label if label else f"Columna_{i + 1}")
    else:
        columns = [f"Columna_{i + 1}" for i in range(len(first_row_clean))]
        # First row is data, not header — yield it
        record = {}
        for i, col in enumerate(columns):
            val = first_row_clean[i] if i < len(first_row_clean) else ""
            record[col] = val
        yield columns, _iter_data_rows(rows_iter, columns, wb, first_record=record)
        return

    yield columns, _iter_data_rows(rows_iter, columns, wb)


def _iter_data_rows(rows_iter, columns, wb, first_record=None):
    """Yields dicts from openpyxl row iterator."""
    try:
        if first_record is not None:
            yield first_record
        for row in rows_iter:
            record = {}
            for i, col in enumerate(columns):
                val = _sanitize_cell(row[i]) if i < len(row) else ""
                record[col] = val
            yield record
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _stream_csv(file_path):
    """Stream rows from CSV using csv module.

    Yields (columns, row_dicts_iterator).
    """
    # Detect encoding
    try:
        with open(file_path, "rb") as f:
            raw = f.read(10000)
        detected = chardet.detect(raw)
        enc = detected.get("encoding") or "latin-1"
    except Exception:
        enc = "latin-1"

    # Try utf-8-sig first, fallback to detected encoding
    for encoding in ["utf-8-sig", enc, "latin-1"]:
        try:
            f = open(file_path, "r", encoding=encoding, newline="")
            reader = csv.reader(f)
            break
        except (UnicodeDecodeError, UnicodeError):
            try:
                f.close()
            except Exception:
                pass
            continue
    else:
        f = open(file_path, "r", encoding="latin-1", newline="", errors="replace")
        reader = csv.reader(f)

    try:
        first_row = next(reader)
    except StopIteration:
        f.close()
        return

    first_row_clean = [c.strip() for c in first_row]
    if _looks_like_header_row(first_row_clean):
        columns = []
        for i, cell in enumerate(first_row_clean):
            label = cell.strip() if cell else ""
            columns.append(label if label else f"Columna_{i + 1}")
    else:
        columns = [f"Columna_{i + 1}" for i in range(len(first_row_clean))]
        record = {}
        for i, col in enumerate(columns):
            record[col] = first_row_clean[i] if i < len(first_row_clean) else ""
        yield columns, _iter_csv_rows(reader, columns, f, first_record=record)
        return

    yield columns, _iter_csv_rows(reader, columns, f)


def _iter_csv_rows(reader, columns, f, first_record=None):
    """Yields dicts from csv reader."""
    try:
        if first_record is not None:
            yield first_record
        for row in reader:
            record = {}
            for i, col in enumerate(columns):
                val = row[i].strip() if i < len(row) else ""
                record[col] = val
            yield record
    finally:
        try:
            f.close()
        except Exception:
            pass


def read_excel(file_path, sheet_name=None):
    """Legacy: returns all rows as list of dicts. USE STREAMING INSTEAD."""
    import pandas as pd
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv":
        df = pd.read_csv(file_path, encoding="utf-8-sig", header=None)
    elif ext in (".xlsx", ".xlsm"):
        df = pd.read_excel(file_path, sheet_name=sheet_name, engine="openpyxl", header=None)
    elif ext == ".xls":
        df = pd.read_excel(file_path, sheet_name=sheet_name, engine="xlrd", header=None)
    else:
        raise ValueError(f"Formato no soportado: {ext}")

    if df is None or df.empty:
        return []

    first_row = df.iloc[0].tolist()
    if _looks_like_header_row(first_row):
        columns = []
        for i, cell in enumerate(first_row):
            label = "" if cell is None else str(cell).strip()
            columns.append(label if label else f"Columna_{i + 1}")
        df.columns = columns
        df = df.iloc[1:]
    else:
        df.columns = [f"Columna_{i + 1}" for i in range(df.shape[1])]

    records = []
    for _, row in df.iterrows():
        record = {}
        for col in df.columns:
            val = row[col]
            record[col] = _sanitize_value(val)
        records.append(record)
    return records
